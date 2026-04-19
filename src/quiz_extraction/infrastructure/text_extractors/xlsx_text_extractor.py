import io

from openpyxl import load_workbook

from src.quiz_extraction.application.ports.text_extractor_port import TextExtractorPort

class XlsxTextExtractor(TextExtractorPort):
    def extract(self, file_bytes: bytes, filename: str) -> str:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None]
                if cells:
                    lines.append(" | ".join(cells))
        return "\n".join(lines).strip()